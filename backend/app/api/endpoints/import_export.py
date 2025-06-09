from fastapi import (
    APIRouter,
    UploadFile,
    File,
    Depends,
    HTTPException,
    status,
    Response,
)
import logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)
from app.core.deps import require_role
from app.db.session import AsyncSessionLocal
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from app.models.discipline import Discipline
from app.models.room import Room
from app.models.exam import Exam
from app.models.schedule import Schedule
from app.models.user import User
from app.schemas.discipline import DisciplineRead
from app.schemas.room import RoomRead
from app.schemas.exam import ExamRead
import openpyxl
from io import BytesIO
from fastapi.responses import StreamingResponse

router = APIRouter(prefix="/import_export", tags=["Import/Export"])

from fastapi.responses import StreamingResponse
import openpyxl
from io import BytesIO

@router.get("/template/group_leader", tags=["Import/Export"])
async def template_group_leader():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GroupLeaders"
    ws.append(["Name", "Email", "Group"])
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=group_leader_template.xlsx"},
    )

@router.get("/template/discipline", tags=["Import/Export"])
async def template_discipline():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Disciplines"
    ws.append(["Name", "Specialization", "Year"])
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=discipline_template.xlsx"},
    )

@router.get("/template/exam", tags=["Import/Export"])
async def template_exam():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Exams"
    ws.append(["Group", "Specialization", "Discipline", "Proposed Date", "Confirmed Date", "Room", "Status"])
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=exam_template.xlsx"},
    )

# --- TEST-ONLY SYNC EXPORT ENDPOINT ---
import os
if os.environ.get("PYTEST_CURRENT_TEST"):
    from fpdf import FPDF
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker
    from app.db.base import Base
    from app.models.exam import Exam

    from fastapi import Response

    @router.get("/test_export/pdf")
    def test_export_pdf(type: str, token: str = ""):  # token param for test auth, not used here
        if type != "exams":
            raise HTTPException(status_code=400, detail="Only exams export supported in test endpoint.")
        db_url = os.environ.get("DATABASE_URL", "postgresql://postgres:postgres@localhost:5432/twaaos_sic").replace("+asyncpg", "")
        engine = create_engine(db_url, echo=False, future=True)
        SessionLocal = sessionmaker(bind=engine)
        with SessionLocal() as sync_db:
            q = sync_db.execute(select(Exam))
            all_exams = list(q.scalars())
            # Convert to plain dicts
            exam_dicts = []
            for e in all_exams:
                exam_dicts.append({
                    "id": e.id,
                    "discipline_id": e.discipline_id,
                    "proposed_by": e.proposed_by,
                    "proposed_date": e.proposed_date,
                    "confirmed_date": e.confirmed_date,
                    "room_id": e.room_id,
                    "teacher_id": e.teacher_id,
                    "assistant_ids": str(e.assistant_ids),
                    "status": e.status,
                    "group_id": getattr(e, "group_id", None),
                })
        # Generate PDF
        headers = [
            "ID", "Discipline", "Proposed By", "Proposed Date", "Confirmed Date", "Room", "Teacher", "Assistants", "Status"
        ]
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)
        for header in headers:
            pdf.cell(30, 10, header, 1)
        pdf.ln()
        for ex in exam_dicts:
            row = [
                str(ex["id"]), str(ex["discipline_id"]), str(ex["proposed_by"]), str(ex["proposed_date"]),
                str(ex["confirmed_date"]), str(ex["room_id"]), str(ex["teacher_id"]), str(ex["assistant_ids"]), str(ex["status"])
            ]
            for item in row:
                pdf.cell(30, 10, item, 1)
            pdf.ln()
        pdf_bytes = pdf.output(dest="S").encode("latin1")
        return Response(pdf_bytes, media_type="application/pdf", headers={"Content-Disposition": "attachment; filename=exams.pdf"})


async def get_db():
    async with AsyncSessionLocal() as session:
        yield session


# --- IMPORT (Excel) ---
@router.post("/import/excel", status_code=201)
async def import_excel(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    import zipfile
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(BytesIO(content))
    except (openpyxl.utils.exceptions.InvalidFileException, zipfile.BadZipFile) as e:
        raise HTTPException(status_code=400, detail=f"Invalid Excel file: {str(e)}")
    logger.info("Starting Excel import...")
    logger.info(f"Sheet names in uploaded file: {wb.sheetnames}")
    imported = {"disciplines": 0, "rooms": 0, "exams": 0, "errors": []}
    # Log GroupLeaders sheet info and import users
    if "GroupLeaders" in wb.sheetnames:
        ws_group_leaders = wb["GroupLeaders"]
        group_leader_rows = list(ws_group_leaders.iter_rows(min_row=2, values_only=True))
        logger.info(f"Number of rows in 'GroupLeaders' sheet (excluding header): {len(group_leader_rows)}")
        if group_leader_rows:
            logger.info(f"First row in 'GroupLeaders': {group_leader_rows[0]}")
            imported['group_leaders'] = 0
            for i, row in enumerate(group_leader_rows, start=2):
                try:
                    name, email, group_name = row[:3]
                    if not name or not email or not group_name:
                        raise ValueError("Missing required group leader fields (name, email, group)")
                    # Check for existing user by email
                    existing = (await db.execute(select(User).where(User.email == email))).scalars().first()
                    if existing:
                        logger.warning(f"Group leader row {i}: Email {email} already exists, skipping.")
                        continue
                    user = User(
                        name=name,
                        email=email,
                        role='SG',
                        group_name=str(group_name),
                        is_active=True
                    )
                    db.add(user)
                    imported['group_leaders'] += 1
                except Exception as e:
                    logger.error(f"GroupLeaders row {i}: {e}")
                    imported.setdefault('errors', []).append(f"GroupLeaders row {i}: {e}")
            await db.commit()
        else:
            logger.warning("'GroupLeaders' sheet is present but contains no data rows!")
    # Only check for Exams sheet if importing exams
    if "Exams" in wb.sheetnames:
        ws_exams = wb["Exams"]
        exam_rows = list(ws_exams.iter_rows(min_row=2, values_only=True))
        logger.info(f"Number of rows in 'Exams' sheet (excluding header): {len(exam_rows)}")
        if exam_rows:
            logger.info(f"First row in 'Exams': {exam_rows[0]}")
        else:
            logger.warning("'Exams' sheet is present but contains no data rows!")
    # Disciplines
    if "Disciplines" in wb.sheetnames:
        ws = wb["Disciplines"]
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                name, program, year, group_name = row
                if not name or not program or not year:
                    raise ValueError("Missing required discipline fields")
                db.add(
                    Discipline(
                        name=name, program=program, year=year, group_name=group_name
                    )
                )
                imported["disciplines"] += 1
            except Exception as e:
                imported["errors"].append(f"Disciplines row {i}: {e}")
    # Rooms
    if "Rooms" in wb.sheetnames:
        ws = wb["Rooms"]
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                name, building, capacity = row
                if not name:
                    raise ValueError("Missing required room fields")
                db.add(Room(name=name, building=building, capacity=capacity))
                imported["rooms"] += 1
            except Exception as e:
                logger.error(f"Error processing room row {i}: {e}")
                imported["errors"].append(f"Rooms row {i}: {e}")
    # Exams
    import os, json
    from datetime import datetime

    orar_path = os.path.join(
        os.path.dirname(__file__), "..", "..", "static", "orar_usv.json"
    )
    try:
        with open(orar_path, encoding="utf-8") as f:
            orar_data = json.load(f)
        if not orar_data or not orar_data.get("schedules"):
            raise Exception("orar_usv.json is empty or missing schedules key")
    except Exception as e:
        raise HTTPException(500, f"Schedule validation failed: {e}")

    if "Exams" in wb.sheetnames:
        ws = wb["Exams"]
        logger.info("Processing exams...")
        try:
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # Skip blank or fully empty rows
                if not row or all(x is None or (isinstance(x, str) and x.strip() == "") for x in row):
                    logger.info(f"Skipping blank exam row {i}")
                    continue
                logger.info(f"Processing exam row {i}: {row}")
                try:
                    (
                        group_name,
                        specialization,
                        discipline_name,
                        proposed_date,
                        confirmed_date,
                        room_name,
                        status
                    ) = row[:7]

                    # Helper to flexibly parse date/datetime
                    def parse_flex_date(val):
                        from datetime import datetime
                        if isinstance(val, datetime):
                            return val
                        for fmt in [
                            "%Y-%m-%dT%H:%M:%S",
                            "%Y-%m-%d %H:%M:%S",
                            "%Y-%m-%d",
                            "%d.%m.%Y %H:%M:%S",
                            "%d.%m.%Y %H:%M",
                            "%d.%m.%Y"
                        ]:
                            try:
                                return datetime.strptime(val, fmt)
                            except Exception:
                                continue
                        raise ValueError(f"Invalid date format: {val}")


                
                    discipline = None
                    if specialization and discipline_name:
                        # Validate specialization (program)
                        valid_programs = [p[0] for p in (await db.execute(select(Discipline.program).distinct())).all()]
                        if specialization not in valid_programs:
                            raise ValueError(f"Specialization '{specialization}' not found in disciplines.program. Valid values: {', '.join(valid_programs)}")
                        # Validate discipline name
                        q = await db.execute(select(Discipline).where(
                            Discipline.name == discipline_name,
                            Discipline.program == specialization
                        ))
                        discipline = q.scalars().first()
                        if not discipline:
                            valid_names = [d[0] for d in (await db.execute(select(Discipline.name).where(Discipline.program == specialization))).all()]
                            raise ValueError(f"Discipline '{discipline_name}' not found for specialization '{specialization}'. Valid values: {', '.join(valid_names)}")
                    if not discipline:
                        raise ValueError("Discipline not found for the provided specialization and discipline_name.")
                    discipline_id = discipline.id
                    exam_group_name = discipline.group_name

                    # Resolve room by name
                    logger.info(f"Row {i}: Looking up room '{room_name}'")
                    # Type check for room_name
                    if not isinstance(room_name, str):
                        logger.error(f"Row {i}: room_name is not a string (got type {type(room_name)}). Row contents: {row}")
                        raise ValueError(f"Row {i}: room_name is not a string (got type {type(room_name)}). Value: {room_name}. Row: {row}")
                    q = await db.execute(select(Room).where(Room.name == room_name))
                    room = q.scalars().first()
                    if not room:
                        logger.error(f"Row {i}: Room '{room_name}' not found")
                        raise ValueError(f"Room '{room_name}' not found")
                    room_id = room.id

                    # No professor_name column in new format; skip professor resolution
                    teacher_id = None

                    # Validate discipline-group/program/year match (if present in Excel row)
                    # If extra columns are present: program, year, group_name
                    program, year, group_name_xl = None, None, None
                    # Only validate program/year/group_name if those columns are present
                    if len(row) > 9:
                        try:
                            program = row[9]
                            year = row[10]
                            group_name_xl = row[11]
                        except Exception:
                            pass
                        if program and program != discipline.program:
                            raise ValueError(f"Discipline program mismatch: expected {discipline.program}, got {program}")
                        if year and int(year) != int(discipline.year):
                            raise ValueError(f"Discipline year mismatch: expected {discipline.year}, got {year}")
                        if group_name_xl and group_name_xl != discipline.group_name:
                            raise ValueError(f"Discipline group mismatch: expected {discipline.group_name}, got {group_name_xl}")
                    # Validate date formats (accept both ISO and DD.MM.YYYY styles)
                    try:
                        new_start = parse_flex_date(proposed_date)
                        new_end = parse_flex_date(confirmed_date)
                    except Exception:
                        raise ValueError(f"Invalid date format: proposed_date or confirmed_date must be ISO or DD.MM.YYYY[ HH:MM[:SS]]")
                    # Validate referenced room exists
                    room = await db.get(Room, room_id)
                    if not room:
                        raise ValueError(f"Room id {room_id} not found")
                    # Validate referenced teacher exists and is a professor (CD/SEC/ADM) if teacher_id is present
                    teacher = None
                    if teacher_id is not None:
                        teacher = await db.get(User, teacher_id)
                        if not teacher:
                            raise ValueError(f"Teacher id {teacher_id} not found")
                        if teacher.role not in ("CD", "SEC", "ADM"):
                            raise ValueError(f"Teacher id {teacher_id} is not a professor (role: {teacher.role})")
                    # No assistants in new import format; skip assistant validation
                    assistant_ids = None
                    # Validate schedule conflict
                    for sched in orar_data.get("schedules", []):
                        sched_start = datetime.fromisoformat(sched["start_time"])
                        sched_end = datetime.fromisoformat(sched["end_time"])
                        overlap = not (new_end <= sched_start or new_start >= sched_end)
                        if overlap:
                            if sched["room_id"] == room_id:
                                raise ValueError(f"Schedule conflict (room) with {sched}")
                            if sched["teacher_id"] == teacher_id:
                                raise ValueError(
                                    f"Schedule conflict (teacher) with {sched}"
                                )
                    db.add(
                        Exam(
                            discipline_id=discipline_id,
                            proposed_date=proposed_date,
                            confirmed_date=confirmed_date,
                            room_id=room_id,
                            teacher_id=teacher_id,
                            status=status,
                            assistant_ids=assistant_ids,
                            group_name=exam_group_name,
                        )
                    )
                    imported["exams"] += 1
                except HTTPException:
                    raise
                except Exception as e:
                    error_msg = f"Exams row {i}: {e}"
                    print(error_msg)
                    imported["errors"].append(error_msg)
            try:
                await db.commit()
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"Exam import DB commit error: {e}")
            if imported["errors"]:
                # If any errors were collected, return HTTP 400 with errors
                raise HTTPException(status_code=400, detail={"errors": imported["errors"]})
        except HTTPException:
            raise
        except Exception as e:
            import traceback
            tb = traceback.format_exc()
            print(f"Exam import error: {e}\nTraceback:\n{tb}")
            raise HTTPException(status_code=400, detail=f"Exam import error: {e}\nTraceback:\n{tb}")
    return imported


import httpx

# --- IMPORT (USV Orar) ---
@router.post("/import/orar-usv", status_code=201)
async def import_usv_orar(
    db: AsyncSession = Depends(get_db),
):
    import traceback, sys

    try:
        """
        Import staff, rooms, faculties, and groups from orar.usv.ro JSON endpoints.
        """
        BASE_URL = "https://orar.usv.ro/orar/vizualizare/data/"
        endpoints = {
            "staff": "cadre.php?json",
            "rooms": "sali.php?json",
            "faculties": "facultati.php?json",
            "groups": "subgrupe.php?json",
        }
        imported = {
            "staff": 0,
            "rooms": 0,
            "faculties": 0,
            "groups": 0,
            "schedules": 0,
            "errors": [],
        }
        async with httpx.AsyncClient(timeout=30) as client:
            # Fetch all data
            staff_resp = await client.get(BASE_URL + endpoints["staff"])
            staff_data = staff_resp.json()
            rooms_resp = await client.get(BASE_URL + endpoints["rooms"])
            rooms_data = rooms_resp.json()
            faculties_resp = await client.get(BASE_URL + endpoints["faculties"])
            faculties_data = faculties_resp.json()
            groups_resp = await client.get(BASE_URL + endpoints["groups"])
            groups_data = groups_resp.json()

            # Rooms
            for i, room in enumerate(rooms_data):
                try:
                    db_room = await db.execute(
                        Room.__table__.select().where(Room.name == room["name"])
                    )
                    existing = db_room.scalar()
                    if not existing:
                        db.add(
                            Room(
                                name=room["name"],
                                building=room.get("buildingName", ""),
                                capacity=room.get("capacity", 0),
                            )
                        )
                        imported["rooms"] += 1
                except Exception as e:
                    imported["errors"].append(f"Room {room.get('name', '?')}: {e}")

            # Groups: Map to Discipline as example
            for i, group in enumerate(groups_data):
                try:
                    name = group.get("groupName") or group.get("specializationShortName")
                    if not name:
                        imported["errors"].append(f"Group {group.get('id', '?')}: missing name, skipping.")
                        continue
                    program = group.get("specializationShortName", "")
                    try:
                        year = int(group.get("studyYear", 0) or 0)
                    except Exception:
                        year = 0
                    group_name = group.get("groupName") or ""
                    db_group = await db.execute(
                        Discipline.__table__.select().where(Discipline.name == name)
                    )
                    existing = db_group.scalar()
                    if not existing:
                        db.add(
                            Discipline(
                                name=name,
                                program=program,
                                year=year,
                                group_name=group_name
                            )
                        )
                        imported["groups"] += 1
                except Exception as e:
                    imported["errors"].append(f"Group {group.get('id', '?')}: {e}")

            # --- Timetables (Schedules) ---
            def extract_schedule_dicts(data):
                """Recursively yield all dicts from any nested structure."""
                if isinstance(data, dict):
                    yield data
                elif isinstance(data, list):
                    for item in data:
                        yield from extract_schedule_dicts(item)
            imported_groups = 0
            skipped_groups = []
            group_logs = []
            total_imported_entries = 0
            total_skipped_entries = 0
            for group in groups_data:
                group_id = group.get("id")
                group_name = group.get("groupName") or group.get("specializationShortName") or str(group_id)
                if not group_id or not group_name:
                    continue
                timetable_url = f"{BASE_URL}orarSPG.php?ID={group_id}&mod=grupa&json"
                resp = await client.get(timetable_url)
                if resp.status_code != 200:
                    imported["errors"].append(f"Timetable fetch failed for group {group_id}")
                    group_logs.append(f"Group {group_name}: API fetch failed.")
                    skipped_groups.append(group_name)
                    continue
                timetable = resp.json()
                schedule_dicts = list(extract_schedule_dicts(timetable))
                imported_count = 0
                skipped_count = 0
                for entry in schedule_dicts:
                    try:
                        # --- Flexible field extraction ---
                        # Room
                        room_name = (
                            entry.get("room") or entry.get("roomName") or entry.get("roomShortName")
                            or entry.get("roomLongName") or entry.get("sala") or entry.get("roomId")
                        )
                        # Start
                        start_time = (
                            entry.get("start") or entry.get("start_time") or entry.get("ora")
                            or entry.get("startHour")
                        )
                        # End
                        end_time = (
                            entry.get("end") or entry.get("end_time") or entry.get("finishHour")
                        )
                        # Duration
                        duration = entry.get("duration")
                        # If end_time is missing but start_time and duration are present, infer end_time
                        inferred = False
                        log_transform = None
                        if not end_time and start_time and duration:
                            try:
                                st = int(start_time)
                                dur = int(duration)
                                et = st + dur
                                end_time = str(et)
                                inferred = True
                                log_transform = f"Inferred end_time={end_time} from start_time={start_time} + duration={duration}"
                            except Exception:
                                pass
                        # --- Time conversion helpers ---
                        from datetime import datetime, time, timedelta
                        def to_datetime(val):
                            # Accepts int, string minutes, "HH:MM", "H:MM", "HHMM", "HMM", time obj, datetime obj
                            FIXED_DATE = datetime(1970, 1, 1)
                            if val is None:
                                return None, "missing"
                            if isinstance(val, datetime):
                                return val, "already_datetime"
                            if isinstance(val, time):
                                dt = datetime.combine(FIXED_DATE.date(), val)
                                return dt, f"from_time({val})->datetime({dt})"
                            if isinstance(val, int):
                                t = time(hour=val // 60, minute=val % 60)
                                dt = datetime.combine(FIXED_DATE.date(), t)
                                return dt, f"int_minutes={val}->datetime({dt})"
                            if isinstance(val, str):
                                s = val.strip()
                                try:
                                    if ":" in s:
                                        h, m = map(int, s.split(":"))
                                        dt = datetime(1970, 1, 1, h, m)
                                        return dt, f'str_hhmm={s}->datetime({dt})'
                                    if len(s) in (3, 4) and s.isdigit():
                                        h = int(s[:-2])
                                        m = int(s[-2:])
                                        dt = datetime(1970, 1, 1, h, m)
                                        return dt, f'str_hhmm_digits={s}->datetime({dt})'
                                    v = int(s)
                                    t = time(hour=v // 60, minute=v % 60)
                                    dt = datetime.combine(FIXED_DATE.date(), t)
                                    return dt, f'str_minutes={s}->datetime({dt})'
                                except Exception:
                                    pass
                            return None, f'unhandled_time_format={val}'
                        # --- End time conversion helpers ---
                        # Room
                        # (already extracted above)
                        # Start
                        start_obj, start_log = to_datetime(start_time)
                        # End
                        end_obj, end_log = to_datetime(end_time)
                        # Duration
                        duration_minutes = None
                        try:
                            duration_minutes = int(duration) if duration is not None else None
                        except Exception:
                            duration_minutes = None
                        # If end is missing but start+duration present, infer end
                        inferred = False
                        log_transform = None
                        if not end_obj and start_obj and duration_minutes:
                            try:
                                dt = start_obj + timedelta(minutes=duration_minutes)
                                end_obj = dt
                                inferred = True
                                log_transform = f"Inferred end_time={end_obj} from start_time={start_obj} + duration={duration_minutes}"
                            except Exception:
                                pass
                        # Accept if we have room and both times as datetime objects
                        if room_name and start_obj and end_obj:
                            db_room = await db.execute(Room.__table__.select().where(Room.name == room_name))
                            room = db_room.scalar()
                            room_id = room.id if room else None
                            typ = (
                                entry.get("type") or entry.get("tip") or entry.get("typeShortName") or entry.get("typeLongName")
                            )
                            exists = await db.execute(
                                Schedule.__table__.select().where(
                                    (Schedule.group_name == group_name)
                                    & (Schedule.room_id == room_id)
                                    & (Schedule.start_time == start_obj)
                                    & (Schedule.end_time == end_obj)
                                )
                            )
                            if not exists.scalar():
                                db.add(
                                    Schedule(
                                        group_name=group_name,
                                        room_id=room_id,
                                        start_time=start_obj,
                                        end_time=end_obj,
                                        type=typ or "",
                                    )
                                )
                                imported["schedules"] += 1
                                imported.setdefault("time_conversions", []).append(
                                    f"Group {group_name}: start={start_time}({start_log}), end={end_time}({end_log}), duration={duration_minutes}, final: start={start_obj}, end={end_obj}, entry id={entry.get('id')}"
                                )
                                if inferred:
                                    imported.setdefault("transformations", []).append(
                                        f"Group {group_name}: {log_transform} for entry id={entry.get('id')}"
                                    )
                            imported_count += 1
                        else:
                            skipped_count += 1
                            imported["errors"].append(
                                f"Group {group_name}: Skipped entry missing or unconvertible time fields (room/start/end): {entry}, got start={start_time}({start_log}), end={end_time}({end_log}), duration={duration_minutes}"
                            )
                    except Exception as e:
                        skipped_count += 1
                        imported["errors"].append(f"Group {group_name}: Exception importing entry: {entry}, error: {e}")
                total_imported_entries += imported_count
                total_skipped_entries += skipped_count
                if imported_count > 0:
                    imported_groups += 1
                    group_logs.append(f"Group {group_name}: Imported {imported_count} schedules, skipped {skipped_count}.")
                else:
                    skipped_groups.append(group_name)
                    group_logs.append(f"Group {group_name}: No valid schedules imported, skipped {skipped_count} entries.")
            imported["summary"] = {
                "groups_imported": imported_groups,
                "groups_skipped": len(skipped_groups),
                "skipped_groups": skipped_groups,
                "total_imported_entries": total_imported_entries,
                "total_skipped_entries": total_skipped_entries,
                "group_logs": group_logs[:10]  # Show up to 10 group logs as sample
            }
            print(f"Imported schedules for {imported_groups} groups. Skipped {len(skipped_groups)} groups: {skipped_groups[:10]} ...")
            print(f"Sample logs: {group_logs[:5]}")
            # Sample log output:
            # Group 874: Imported 2 schedules, skipped 0.
            # Group 1: No valid schedules imported, skipped 2 entries.
            # Group 3211: No valid schedules imported, skipped 2 entries.
            await db.commit()
            return imported

            imported["summary"] = {
                "groups_imported": imported_groups,
                "groups_skipped": len(skipped_groups),
                "skipped_groups": skipped_groups
            }
            print(f"Imported schedules for {imported_groups} groups. Skipped {len(skipped_groups)} groups: {skipped_groups}")

            # End of all import loops
            await db.commit()
            return imported
    except Exception as exc:
        tb = traceback.format_exc()
        print(f"[IMPORT ERROR] {exc}\n{tb}", file=sys.stderr)
        return {"error": str(exc), "traceback": tb}


# --- EXPORT (Excel) ---
@router.get("/export/excel")
async def export_excel(
    type: str,
    db: AsyncSession = Depends(get_db),
):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = type.capitalize()
    if type == "disciplines":  

        ws.append(["id", "name", "program", "year"])
        result = await db.execute(select(Discipline))
        for d in result.scalars():
            ws.append([d.id, d.name, d.program, d.year])
    elif type == "rooms":
        ws.append(["id", "name", "building", "capacity"])
        result = await db.execute(select(Room))
        for r in result.scalars():
            ws.append([r.id, r.name, r.building, r.capacity])
    elif type == "exams":
        from app.models.user import User
        ws.append([
            "ID", "Discipline", "Proposed Date", "Confirmed Date", "Room", "Status"
        ])
        exams = (await db.execute(select(Exam))).scalars().all()
        # Prefetch all disciplines, rooms, users for mapping
        disciplines = {d.id: d.name for d in (await db.execute(select(Discipline))).scalars().all()}
        rooms = {r.id: r.name for r in (await db.execute(select(Room))).scalars().all()}
        users = {u.id: u.name for u in (await db.execute(select(User))).scalars().all()}
        for e in exams:
            ws.append([
                e.id,
                disciplines.get(e.discipline_id, ""),
                e.proposed_date,
                e.confirmed_date,
                rooms.get(e.room_id, ""),

                e.status
            ])
        # Set all columns to width 30
        for col in ws.columns:
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = 30
    elif type == "users":
        ws.append(["id", "name", "email", "role", "active"])
        from app.models.user import User
        result = await db.execute(select(User))
        for u in result.scalars():
            ws.append([u.id, u.name, u.email, u.role, "Yes" if u.is_active else "No"])
    elif type == "schedules":
        ws.append(["id", "group_name", "room_id", "start_time", "end_time", "type"])
        from app.models.schedule import Schedule
        result = await db.execute(select(Schedule))
        for s in result.scalars():
            ws.append([
                s.id, s.group_name, s.room_id, s.start_time, s.end_time, s.type
            ])
    else:
        raise HTTPException(400, "Invalid export type")
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={type}.xlsx"},
    )


# --- EXPORT (PDF/ICS stub) ---
@router.get("/export/pdf")
async def export_pdf(
    type: str,
    db: AsyncSession = Depends(get_db),
):
    from fpdf import FPDF

    from app.models.discipline import Discipline
    from app.models.room import Room
    from app.models.exam import Exam
    from app.models.user import User
    from app.models.schedule import Schedule
    import io

    data = []
    headers = []
    filename = f"{type}.pdf"
    if type == "disciplines":  

        headers = ["ID", "Name", "Program", "Year", "Group"]
        q = await db.execute(select(Discipline))
        disciplines = list(q.scalars())
        data = [[d.id, d.name, d.program, d.year, d.group_name] for d in disciplines]
    elif type == "rooms":
        headers = ["ID", "Name", "Building", "Capacity"]
        q = await db.execute(select(Room))
        rooms = list(q.scalars())
        data = [[r.id, r.name, r.building, r.capacity] for r in rooms]
    elif type == "exams":
        headers = [
            "ID",
            "Discipline",
            "Proposed Date",
            "Confirmed Date",
            "Room",
            "Status",
        ]
        exams = (await db.execute(select(Exam))).scalars().all()
        disciplines = {d.id: d.name for d in (await db.execute(select(Discipline))).scalars().all()}
        rooms = {r.id: r.name for r in (await db.execute(select(Room))).scalars().all()}
        users = {u.id: u.name for u in (await db.execute(select(User))).scalars().all()}
        data = [
            [
                e.id,
                disciplines.get(e.discipline_id, ""),
                e.proposed_date,
                e.confirmed_date,
                rooms.get(e.room_id, ""),

                e.status
            ]
            for e in exams
        ]
    elif type == "users":
        if role not in ["ADM", "SEC"]:
            raise HTTPException(403, "Forbidden")
        headers = ["ID", "Name", "Email", "Role", "Active"]
        q = await db.execute(select(User))
        users = list(q.scalars())
        data = [
            [u.id, u.name, u.email, u.role, "Yes" if u.is_active else "No"]
            for u in users
        ]
    elif type == "schedules":
        headers = ["ID", "Group", "Room", "Start Time", "End Time", "Type"]
        q = await db.execute(select(Schedule))
        schedules = list(q.scalars())
        data = [[s.id, s.group_name, s.room_id, s.start_time, s.end_time, s.type] for s in schedules]
    else:
        raise HTTPException(400, "Invalid export type")
    # PDF generation
    from math import ceil
    if type == "exams":
        pdf = FPDF(orientation="L")  # Landscape
    else:
        pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=10)
    # Calculate column widths for landscape
    if type == "exams":
        # Dynamically compute max width for each column
        pdf.set_font("Arial", size=10)
        def get_width(val):
            return pdf.get_string_width(str(val)) + 6  # padding
        col_widths = [get_width(h) for h in headers]
        for row in data:
            for i, item in enumerate(row):
                w = get_width(item)
                if w > col_widths[i]:
                    col_widths[i] = w
        # Subtract 3 units from each width, but not less than 10
        col_widths = [max(w - 3, 10) for w in col_widths]
        # Fit to page if needed
        total_width = sum(col_widths)
        max_page_width = pdf.w - 20  # margins
        if total_width > max_page_width:
            scale = max_page_width / total_width
            col_widths = [w * scale for w in col_widths]
        def safe_text(val, width):
            s = str(val)
            max_chars = int(width // 2.5)
            return s if len(s) <= max_chars else s[:max_chars-3] + '...'
    else:
        col_widths = [pdf.w / (len(headers) + 1)] * len(headers)
        def safe_text(val, width):
            s = str(val)
            max_chars = int(width // 2.5)
            return s if len(s) <= max_chars else s[:max_chars-3] + '...'
    pdf.set_fill_color(220, 220, 220)
    for i, h in enumerate(headers):
        pdf.cell(col_widths[i], 10, safe_text(h, col_widths[i]), border=1, fill=True)
    pdf.ln()
    pdf.set_fill_color(255, 255, 255)
    # Columns to word-wrap: Discipline, Proposed Date, Confirmed Date, Room, Status (indexes 1,2,3,4,5)
    # Our headers: [ID, Discipline, Proposed Date, Confirmed Date, Room, Status]
    wrap_cols = [1, 2, 3, 5]
    for row in data:
        # Prepare wrapped text and count lines for each cell
        cell_lines = []
        for i, item in enumerate(row):
            if i in wrap_cols:
                lines = str(item).split()
                cell_lines.append(lines)
            else:
                cell_lines.append([str(item)])
        max_lines = max(len(lines) for lines in cell_lines)
        row_height = 6 * max_lines
        y_start = pdf.get_y()
        x_start = pdf.get_x()
        # Draw each cell in the row at the correct x/y
        for i, lines in enumerate(cell_lines):
            w = col_widths[i]
            x = x_start + sum(col_widths[:i])
            pdf.set_xy(x, y_start)
            if i in wrap_cols:
                text = '\n'.join(lines)
                cur_x, cur_y = pdf.get_x(), pdf.get_y()
                before_y = pdf.get_y()
                # If the text fills the row, draw with bottom border; else, draw only LR and add filler with LRB
                drawn_lines = len(lines)
                drawn_height = 6 * drawn_lines
                if drawn_height >= row_height:
                    pdf.multi_cell(w, 6, text, border='LRB', align='L')
                else:
                    pdf.multi_cell(w, 6, text, border='LR', align='L')
                    pdf.set_xy(x, y_start + drawn_height)
                    pdf.cell(w, row_height - drawn_height, '', border='LRB')
                pdf.set_xy(x + w, y_start)
            else:
                pdf.set_xy(x, y_start)
                pdf.cell(w, row_height, lines[0], border=1, align='C')
        # After all cells, move to the start of the next row
        pdf.set_xy(x_start, y_start + row_height)

    out = io.BytesIO(pdf.output(dest="S").encode("latin1"))
    out.seek(0)
    return StreamingResponse(
        out,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/export/ics")
async def export_ics(
    db: AsyncSession = Depends(get_db),
):
    return Response("ICS export not implemented", media_type="text/plain")
