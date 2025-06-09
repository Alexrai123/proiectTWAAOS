import pytest
from httpx import AsyncClient
from jose import jwt
from app.core.deps import SECRET_KEY, ALGORITHM
import io
import openpyxl

# Helper to make JWT tokens for different roles
ROLES = [
    ("ADM", jwt.encode({"sub": str(1), "role": "ADM"}, SECRET_KEY, algorithm=ALGORITHM)),
    ("SEC", jwt.encode({"sub": str(2), "role": "SEC"}, SECRET_KEY, algorithm=ALGORITHM)),
    ("CD", jwt.encode({"sub": str(3), "role": "CD"}, SECRET_KEY, algorithm=ALGORITHM)),
    ("SG", jwt.encode({"sub": str(4), "role": "SG"}, SECRET_KEY, algorithm=ALGORITHM)),
]

@pytest.mark.asyncio
@pytest.mark.parametrize("role,token", ROLES)
async def test_import_excel_permissions(role, token, async_client):
    # Create required related objects first
    headers = {"Authorization": f"Bearer {token}"}

    if role == "ADM":
        # All setup and import as ADM
        adm_headers = {"Authorization": f"Bearer {token}"}
        # Discipline
        disc_resp = await async_client.post("/disciplines/", json={"name": "Math", "program": "SCI", "year": 1, "group_name": "A1"}, headers=adm_headers)
        assert disc_resp.status_code in (200, 201), f"Discipline creation failed: {disc_resp.text}"
        disc_id = disc_resp.json()["id"]
        # Room
        room_resp = await async_client.post("/rooms/", json={"name": "101", "building": "Main", "capacity": 30}, headers=adm_headers)
        assert room_resp.status_code in (200, 201), f"Room creation failed: {room_resp.text}"
        room_id = room_resp.json()["id"]
        # Teacher
        teacher_payload = {"username": "teach1", "password": "pw", "role": "CD", "full_name": "Teach One", "name": "Teach One", "email": "teach1@example.com"}
        teacher_resp = await async_client.post("/users/", json=teacher_payload, headers=adm_headers)
        assert teacher_resp.status_code in (200, 201), f"Teacher creation failed: {teacher_resp.text}"
        teacher_id = teacher_resp.json()["id"]




        # Create Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Exams"
        ws.append(["discipline_id", "proposed_by", "proposed_date", "confirmed_date", "room_id", "teacher_id", "status"])
        ws.append([
            disc_id,
            teacher_id,
            "2025-06-07T10:00:00",
            "2025-06-07T12:00:00",
            room_id,
            teacher_id,
            "scheduled"
        ])
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        files = {"file": ("exams.xlsx", stream.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        resp = await async_client.post("/import_export/import/excel", files=files, headers=adm_headers)
        print("\nRESPONSE STATUS:", resp.status_code)
        print("RESPONSE BODY:", resp.text)
        assert resp.status_code in (200, 201)
        data = resp.json()
        assert data.get("exams", 0) >= 1
    elif role == "SEC":
        # Setup as ADM, import as SEC
        adm_token = ROLES[0][1]
        adm_headers = {"Authorization": f"Bearer {adm_token}"}
        # Discipline
        disc_resp = await async_client.post("/disciplines/", json={"name": "Math", "program": "SCI", "year": 1, "group_name": "A1"}, headers=adm_headers)
        assert disc_resp.status_code in (200, 201), f"Discipline creation failed: {disc_resp.text}"
        disc_id = disc_resp.json()["id"]
        # Room
        room_resp = await async_client.post("/rooms/", json={"name": "101", "building": "Main", "capacity": 30}, headers=adm_headers)
        assert room_resp.status_code in (200, 201), f"Room creation failed: {room_resp.text}"
        room_id = room_resp.json()["id"]
        # Teacher
        teacher_payload = {"username": "teach1", "password": "pw", "role": "CD", "full_name": "Teach One", "name": "Teach One", "email": "teach1@example.com"}
        teacher_resp = await async_client.post("/users/", json=teacher_payload, headers=adm_headers)
        assert teacher_resp.status_code in (200, 201), f"Teacher creation failed: {teacher_resp.text}"
        teacher_id = teacher_resp.json()["id"]




        # Create Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Exams"
        ws.append(["discipline_id", "proposed_by", "proposed_date", "confirmed_date", "room_id", "teacher_id", "status"])
        ws.append([
            disc_id,
            teacher_id,
            "2025-06-07T10:00:00",
            "2025-06-07T12:00:00",
            room_id,
            teacher_id,
            "scheduled"
        ])
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        files = {"file": ("exams.xlsx", stream.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        sec_headers = {"Authorization": f"Bearer {token}"}
        resp = await async_client.post("/import_export/import/excel", files=files, headers=sec_headers)
        print("\nRESPONSE STATUS:", resp.status_code)
        print("RESPONSE BODY:", resp.text)
        assert resp.status_code in (200, 201)
        data = resp.json()
        assert data.get("exams", 0) >= 1
    else:
        import uuid
        from jose import jwt
        from app.core.deps import SECRET_KEY, ALGORITHM
        forbidden_group = "OtherGroup"
        adm_token = ROLES[0][1]
        adm_headers = {"Authorization": f"Bearer {adm_token}"}
        # --- SG forbidden group test ---
        if role == "SG":
            # Create SG user with group_name 'A1'
            sg_payload = {"username": f"sg_{uuid.uuid4()}", "password": "pw", "role": "SG", "full_name": "SG User", "name": "SG User", "email": f"sg_{uuid.uuid4()}@example.com", "group_name": "A1"}
            sg_resp = await async_client.post("/users/", json=sg_payload, headers=adm_headers)
            assert sg_resp.status_code in (200, 201), f"SG creation failed: {sg_resp.text}"
            sg_id = sg_resp.json()["id"]
            sg_token = jwt.encode({"sub": str(sg_id), "role": "SG", "group_name": "A1"}, SECRET_KEY, algorithm=ALGORITHM)
            sg_headers = {"Authorization": f"Bearer {sg_token}"}
            # Discipline with a different group
            disc_resp = await async_client.post("/disciplines/", json={"name": "Math", "program": "SCI", "year": 1, "group_name": forbidden_group}, headers=adm_headers)
            assert disc_resp.status_code in (200, 201), f"Discipline creation failed: {disc_resp.text}"
            disc_id = disc_resp.json()["id"]
            # Room
            room_resp = await async_client.post("/rooms/", json={"name": "101", "building": "Main", "capacity": 30}, headers=adm_headers)
            assert room_resp.status_code in (200, 201), f"Room creation failed: {room_resp.text}"
            room_id = room_resp.json()["id"]
            # Teacher
            teacher_payload = {"username": f"teach_{uuid.uuid4()}", "password": "pw", "role": "CD", "full_name": "Teach One", "name": "Teach One", "email": f"teach_{uuid.uuid4()}@example.com"}
            teacher_resp = await async_client.post("/users/", json=teacher_payload, headers=adm_headers)
            assert teacher_resp.status_code in (200, 201), f"Teacher creation failed: {teacher_resp.text}"
            teacher_id = teacher_resp.json()["id"]
            # Create Excel
    
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Exams"
            ws.append(["discipline_id", "proposed_by", "proposed_date", "confirmed_date", "room_id", "teacher_id", "status", "group_name"])
            ws.append([
                disc_id,
                teacher_id,
                "2025-06-07T10:00:00",
                "2025-06-07T12:00:00",
                room_id,
                teacher_id,
                "scheduled",
                forbidden_group
            ])
            stream = io.BytesIO()
            wb.save(stream)
            stream.seek(0)
            files = {"file": ("exams.xlsx", stream.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
            # Import as SG (forbidden group)
            resp = await async_client.post("/import_export/import/excel", files=files, headers=sg_headers)
            assert resp.status_code == 403
        else:
            # For CD, just check forbidden logic or skip (no group-based restriction)
            pass



@pytest.mark.asyncio
async def test_import_excel_bad_format(async_client):
    # Send a non-Excel file
    files = {"file": ("not_excel.txt", b"not an excel file", "text/plain")}
    # Use ADM token
    token = jwt.encode({"sub": str(1), "role": "ADM"}, SECRET_KEY, algorithm=ALGORITHM)
    headers = {"Authorization": f"Bearer {token}"}
    resp = await async_client.post("/import_export/import/excel", files=files, headers=headers)
    assert resp.status_code in (400, 422)
    # Optionally check error message
    # assert "error" in resp.text or resp.json().get("errors")
